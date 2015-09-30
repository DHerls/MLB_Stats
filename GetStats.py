from urllib.request import urlopen
import json
import random
import openpyxl
import csv

# URL is split into parts for easy editing of what data is collected
# Base URL where the player repository is located
URLBase = "http://mlb.mlb.com/pubajax/wf/flow/stats.splayer"
# What season the data is targeting
URLSeason = "?season=2015"
# The order that the players appear on the page, entirely useless for this program
URLOrder = "&sort_order=%27desc%27"
# By what stat the players are being sorted, again useless
URLColumn = "&sort_column=%27avg%27"
# Hitting data as opposed to fielding data
URLType = "&stat_type=hitting"
# Used by the Stats page to request a certain type of webpage, useless
URLPage = "&page_type=SortablePlayer"
# Regular Season Games
URLGame = "&game_type=%27R%27"
# All players in the league
URLPlayer = "&player_pool=ALL"
# Single season as opposed to all time
URLSeasonType = "&season_type=ANY"
# AL vs NL vs MLB results
URLSportCode = "&sport_code=%27mlb%27"
# Unknown
URLResults = "&results=1000"
# Which page of results should be loaded
URLPageNum = "&recSP=1"
# Loads 1250 players on one page which is more than the league has, ensuring that every player appears on one page
URLResultNum = "&recPP=1250"


# Combines the URL fragments above into a single string
def get_url():
    return URLBase + URLSeason + URLOrder + URLColumn + URLType + URLPage + URLGame + URLSeasonType + URLSportCode + \
           URLResults + URLPageNum + URLResultNum

# Sends an HTTP request to the MLB stats website and loads that data into a format the computer can read
siteData = json.loads(urlopen(get_url()).read().decode("UTF-8"))
# Gets list of player data from raw site data
playerList = siteData["stats_sortable_player"]["queryResults"]["row"]


listToDelete = []

# If player has <20 at bats or is a pitcher, they are marked for removal
for player in playerList:
    if int(player.get("ab")) < 20:
        listToDelete.append(player)
    else:
        if player.get("pos") == "P":
            listToDelete.append(player)

# The players marked for removal are removed
for player in listToDelete:
    playerList.remove(player)

# Gets the number of players left in the list
length = len(playerList)

randomNumbers = []

# Generates a list of 40 random numbers
while len(randomNumbers) < 40:
    # Generates a random number from 0 to the number of players in the eligible players list -1 because python is 0 base
    rand = random.randint(0, length-1)
    # If the number isn't a duplicate, it stores the number in the list of random numbers
    if not randomNumbers.__contains__(rand):
        randomNumbers.append(rand)

randomPlayers = []

# Uses the random numbers to retrieve the corresponding players
for num in randomNumbers:
    randomPlayers.append(playerList[num])

# Create Excel Workbook
wb = openpyxl.Workbook()
# Get a new Worksheet
dataSheet = wb.get_active_sheet()
# Set Worksheet title to Data
dataSheet.title = "Data"

dataList = ["name_display_first_last", "team_abbrev", "pos", "g", "so", "avg"]

# Put titles for the columns
dataSheet['A1'] = "Name"
dataSheet['B1'] = "Team"
dataSheet['C1'] = "Position"
dataSheet['D1'] = "Games"
dataSheet['E1'] = "Strikeouts"
dataSheet['F1'] = "Average"

# Put data in the Excel workbook
row = 2
for player in randomPlayers:
    column = 'A'
    for key in dataList:
        value = player.get(key)
        # See if the value is an integer
        try:
            value = int(value)
        except ValueError:
            # See if the value is a decimal
            try:
                value = float(value)
            except ValueError:
                False

        dataSheet[column + str(row)] = value
        column = chr(ord(column)+1)
    row += 1

# Create a worksheet called "Frequency"
frequencySheet = wb.create_sheet()
frequencySheet.title = 'Frequency'

# Reads the accompanying formula sheet and pastes them into the appropriate cells
for key, val in csv.reader(open("Frequency.csv")):
    frequencySheet[key] = val

# Makes some cells look a little bit pretty
frequencySheet.merge_cells("A1:G1")
frequencySheet.merge_cells("A11:G11")
frequencySheet.merge_cells("A21:G21")

# Saves the workbook
wb.save("results.xlsx")
